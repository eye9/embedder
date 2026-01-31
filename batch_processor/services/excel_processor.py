"""Excel file processing utilities for batch processor."""

import pandas as pd
import logging
from pathlib import Path
from typing import Iterator, Tuple, Optional, List, Dict, Any
import time
from dataclasses import dataclass

logger = logging.getLogger(__name__)


@dataclass
class ValidationResult:
    """Result of Excel file validation."""
    is_valid: bool
    error_message: str
    total_rows: int
    has_description_column: bool = False
    has_hts_code_column: bool = False
    has_url_column: bool = False
    url_column_name: Optional[str] = None
    empty_descriptions: int = 0
    existing_hts_codes: int = 0


class ExcelProcessor:
    """Handles Excel file processing with memory-efficient chunked reading."""
    
    def __init__(self, chunk_size: int = 1000):
        """
        Initialize ExcelProcessor.
        
        Args:
            chunk_size: Number of rows to process in each chunk for memory efficiency
        """
        self.chunk_size = chunk_size
        self.required_columns = ["Product Detailed Description"]
        self.optional_columns = ["HTS Code"]
        self.url_column_names = [
            "Link to customer's web-page with item description",
            "URL",
            "Product URL",
            "Link"
        ]
        
        logger.info(f"ExcelProcessor initialized with chunk_size={chunk_size}")
    
    def get_file_info(self, file_path: Path) -> Dict[str, Any]:
        """
        Get detailed information about Excel file structure and content.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with file information
        """
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            
            # Find columns
            description_col = next(
                (col for col in df.columns if "Product Detailed Description" in str(col)), 
                None
            )
            
            hts_col = None
            for col in df.columns:
                if "HTS Code" in str(col) or "HTS_Code" in str(col):
                    hts_col = col
                    break
            
            # Find URL column
            url_col = self._find_url_column(df.columns)
            
            # Count rows with descriptions
            rows_with_descriptions = 0
            if description_col:
                rows_with_descriptions = (~df[description_col].isna()).sum() - \
                                       (df[description_col].astype(str).str.strip() == '').sum()
            
            # Count rows with existing codes
            rows_with_existing_codes = 0
            if hts_col:
                rows_with_existing_codes = (~df[hts_col].isna()).sum() - \
                                         (df[hts_col].astype(str).str.strip() == '').sum()
            
            # Count rows with URLs
            rows_with_urls = 0
            if url_col:
                rows_with_urls = (~df[url_col].isna()).sum() - \
                               (df[url_col].astype(str).str.strip() == '').sum()
            
            return {
                "total_rows": len(df),
                "rows_with_descriptions": rows_with_descriptions,
                "rows_with_existing_codes": rows_with_existing_codes,
                "rows_with_urls": rows_with_urls,
                "columns": df.columns.tolist(),
                "has_description_column": description_col is not None,
                "has_hts_column": hts_col is not None,
                "has_url_column": url_col is not None,
                "description_column": description_col,
                "hts_column": hts_col,
                "url_column": url_col
            }
            
        except Exception as e:
            logger.error(f"Failed to get file info: {e}")
            return {
                "total_rows": 0,
                "rows_with_descriptions": 0,
                "rows_with_existing_codes": 0,
                "rows_with_urls": 0,
                "columns": [],
                "has_description_column": False,
                "has_hts_column": False,
                "has_url_column": False,
                "description_column": None,
                "hts_column": None,
                "url_column": None
            }

    def validate_file(self, file_path: Path) -> Tuple[bool, str, int]:
        """
        Validate Excel file format and required columns.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple of (is_valid, error_message, total_rows)
        """
        logger.info(f"Validating Excel file: {file_path}")
        
        try:
            # Check if file exists
            if not file_path.exists():
                return False, f"File not found: {file_path}", 0
            
            # Check file extension
            if file_path.suffix.lower() not in ['.xlsx', '.xls']:
                return False, f"Invalid file format. Expected .xlsx or .xls, got {file_path.suffix}", 0
            
            # Try to read the file header to check structure
            try:
                # Read just the first few rows to check columns
                df_sample = pd.read_excel(file_path, nrows=5, engine='openpyxl')
                
                if df_sample.empty:
                    return False, "File is empty or contains no data", 0
                
                # Check for required columns
                columns = df_sample.columns.tolist()
                has_description = any(
                    col for col in columns 
                    if "Product Detailed Description" in str(col)
                )
                
                if not has_description:
                    return False, f"Required column 'Product Detailed Description' not found. Available columns: {columns}", len(df_sample)
                
                # Check for optional HTS Code column
                has_hts_code = any(
                    col for col in columns 
                    if "HTS Code" in str(col) or "HTS_Code" in str(col)
                )
                
                # Check for URL column
                url_column = self._find_url_column(columns)
                has_url_column = url_column is not None
                
                # Get total row count efficiently
                df_full = pd.read_excel(file_path, engine='openpyxl')
                total_rows = len(df_full)
                
                # Count empty descriptions and existing HTS codes
                description_col = next(
                    col for col in df_full.columns 
                    if "Product Detailed Description" in str(col)
                )
                
                empty_descriptions = df_full[description_col].isna().sum() + \
                                   (df_full[description_col].astype(str).str.strip() == '').sum()
                
                existing_hts_codes = 0
                if has_hts_code:
                    hts_col = next(
                        (col for col in df_full.columns 
                         if "HTS Code" in str(col) or "HTS_Code" in str(col)), 
                        None
                    )
                    if hts_col:
                        existing_hts_codes = (~df_full[hts_col].isna()).sum() - \
                                           (df_full[hts_col].astype(str).str.strip() == '').sum()
                
                logger.info(f"File validation successful: {total_rows} rows, "
                           f"{empty_descriptions} empty descriptions, "
                           f"{existing_hts_codes} existing HTS codes, "
                           f"URL column: {url_column}")
                
                return True, "", total_rows
                
            except Exception as e:
                return False, f"Failed to read Excel file: {str(e)}", 0
                
        except Exception as e:
            logger.error(f"File validation failed: {e}")
            return False, f"Validation error: {str(e)}", 0
    
    def read_file_chunked(
        self, 
        file_path: Path, 
        process_mode: str = "all"
    ) -> Iterator[Tuple[pd.DataFrame, int, int]]:
        """
        Read Excel file in chunks for memory-efficient processing.
        
        Args:
            file_path: Path to the Excel file
            process_mode: "all" to process all rows, "empty_only" to process only rows without HTS codes
            
        Yields:
            Tuple of (chunk_dataframe, chunk_start_row, total_rows)
        """
        logger.info(f"Reading file in chunks: {file_path}, mode: {process_mode}")
        
        try:
            # First, get the total row count
            df_full = pd.read_excel(file_path, engine='openpyxl')
            total_rows = len(df_full)
            
            # Find the description column
            description_col = next(
                col for col in df_full.columns 
                if "Product Detailed Description" in str(col)
            )
            
            # Find HTS Code column if it exists
            hts_col = None
            for col in df_full.columns:
                if "HTS Code" in str(col) or "HTS_Code" in str(col):
                    hts_col = col
                    break
            
            # Process in chunks
            for start_row in range(0, total_rows, self.chunk_size):
                end_row = min(start_row + self.chunk_size, total_rows)
                chunk = df_full.iloc[start_row:end_row].copy()
                
                # Filter chunk based on process_mode
                filtered_chunk = self.filter_rows_for_processing(chunk, process_mode)
                
                if not filtered_chunk.empty:
                    logger.debug(f"Yielding chunk: rows {start_row}-{end_row-1}, "
                               f"{len(filtered_chunk)} rows to process")
                    yield filtered_chunk, start_row, total_rows
                else:
                    logger.debug(f"Skipping empty chunk: rows {start_row}-{end_row-1}")
                    
        except Exception as e:
            logger.error(f"Failed to read file in chunks: {e}")
            raise
    
    def read_file_chunked_with_urls(
        self, 
        file_path: Path, 
        process_mode: str = "all"
    ) -> Iterator[Tuple[pd.DataFrame, int, int, Optional[str]]]:
        """
        Read Excel file in chunks with URL column support for memory-efficient processing.
        
        Args:
            file_path: Path to the Excel file
            process_mode: "all" to process all rows, "empty_only" to process only rows without HTS codes
            
        Yields:
            Tuple of (chunk_dataframe, chunk_start_row, total_rows, url_column_name)
        """
        logger.info(f"Reading file in chunks with URL support: {file_path}, mode: {process_mode}")
        
        try:
            # First, get the total row count and identify URL column
            df_full = pd.read_excel(file_path, engine='openpyxl')
            total_rows = len(df_full)
            url_column = self._find_url_column(df_full.columns)
            
            # Find the description column
            description_col = next(
                col for col in df_full.columns 
                if "Product Detailed Description" in str(col)
            )
            
            # Find HTS Code column if it exists
            hts_col = None
            for col in df_full.columns:
                if "HTS Code" in str(col) or "HTS_Code" in str(col):
                    hts_col = col
                    break
            
            logger.info(f"URL column detected: {url_column}")
            
            # Process in chunks
            for start_row in range(0, total_rows, self.chunk_size):
                end_row = min(start_row + self.chunk_size, total_rows)
                chunk = df_full.iloc[start_row:end_row].copy()
                
                # Filter chunk based on process_mode
                filtered_chunk = self.filter_rows_for_processing(chunk, process_mode)
                
                if not filtered_chunk.empty:
                    logger.debug(f"Yielding chunk: rows {start_row}-{end_row-1}, "
                               f"{len(filtered_chunk)} rows to process, URL column: {url_column}")
                    yield filtered_chunk, start_row, total_rows, url_column
                else:
                    logger.debug(f"Skipping empty chunk: rows {start_row}-{end_row-1}")
                    
        except Exception as e:
            logger.error(f"Failed to read file in chunks with URLs: {e}")
            raise
    
    def _find_url_column(self, columns: List[str]) -> Optional[str]:
        """
        Find URL column among possible column names.
        
        Args:
            columns: List of column names from DataFrame
            
        Returns:
            URL column name if found, None otherwise
        """
        for col_name in columns:
            if col_name in self.url_column_names:
                return col_name
        return None
    
    def extract_url_from_row(self, row: pd.Series, url_column: Optional[str]) -> Optional[str]:
        """
        Extract URL from a DataFrame row.
        
        Args:
            row: DataFrame row (Series)
            url_column: Name of the URL column
            
        Returns:
            URL string if found and valid, None otherwise
        """
        if not url_column or url_column not in row.index:
            return None
        
        url_value = row[url_column]
        if pd.isna(url_value):
            return None
        
        url_str = str(url_value).strip()
        return url_str if url_str else None
    
    def validate_file_with_url_support(self, file_path: Path) -> Tuple[bool, str, int, bool]:
        """
        Validate Excel file with URL column detection support.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple of (is_valid, error_message, total_rows, has_url_column)
        """
        logger.info(f"Validating Excel file with URL support: {file_path}")
        
        try:
            # Check if file exists
            if not file_path.exists():
                return False, f"File not found: {file_path}", 0, False
            
            # Check file extension
            if file_path.suffix.lower() not in ['.xlsx', '.xls']:
                return False, f"Invalid file format. Expected .xlsx or .xls, got {file_path.suffix}", 0, False
            
            # Try to read the file header to check structure
            try:
                # Read just the first few rows to check columns
                df_sample = pd.read_excel(file_path, nrows=5, engine='openpyxl')
                
                if df_sample.empty:
                    return False, "File is empty or contains no data", 0, False
                
                # Check for required columns
                columns = df_sample.columns.tolist()
                has_description = any(
                    col for col in columns 
                    if "Product Detailed Description" in str(col)
                )
                
                if not has_description:
                    return False, f"Required column 'Product Detailed Description' not found. Available columns: {columns}", len(df_sample), False
                
                # Check for URL column
                url_column = self._find_url_column(columns)
                has_url_column = url_column is not None
                
                # Get total row count efficiently
                df_full = pd.read_excel(file_path, engine='openpyxl')
                total_rows = len(df_full)
                
                logger.info(f"File validation successful: {total_rows} rows, "
                           f"URL column: {url_column if has_url_column else 'Not found'}")
                
                return True, "", total_rows, has_url_column
                
            except Exception as e:
                return False, f"Failed to read Excel file: {str(e)}", 0, False
                
        except Exception as e:
            logger.error(f"File validation failed: {e}")
            return False, f"Validation error: {str(e)}", 0, False
    
    def filter_rows_for_processing(
        self, 
        df: pd.DataFrame, 
        process_mode: str
    ) -> pd.DataFrame:
        """
        Filter rows based on processing mode.
        
        Args:
            df: DataFrame chunk to filter
            process_mode: "all" or "empty_only"
            
        Returns:
            Filtered DataFrame
        """
        if process_mode not in ["all", "empty_only"]:
            raise ValueError(f"Invalid process_mode: {process_mode}. Must be 'all' or 'empty_only'")
        
        # Find the description column
        description_col = next(
            (col for col in df.columns if "Product Detailed Description" in str(col)), 
            None
        )
        
        if not description_col:
            logger.warning("No 'Product Detailed Description' column found")
            return pd.DataFrame()
        
        # Always filter out rows with empty descriptions
        mask = df[description_col].notna() & (df[description_col].astype(str).str.strip() != '')
        
        if process_mode == "empty_only":
            # Additionally filter out rows that already have HTS codes
            hts_col = None
            for col in df.columns:
                if "HTS Code" in str(col) or "HTS_Code" in str(col):
                    hts_col = col
                    break
            
            if hts_col:
                # Only process rows where HTS code is empty/null
                hts_mask = df[hts_col].isna() | (df[hts_col].astype(str).str.strip() == '')
                mask = mask & hts_mask
                
                logger.debug(f"Empty-only mode: filtering {len(df)} rows to {mask.sum()} rows")
            else:
                logger.info("No HTS Code column found, processing all rows with descriptions")
        
        filtered_df = df[mask].copy()
        logger.debug(f"Filtered {len(df)} rows to {len(filtered_df)} rows for processing")
        
        return filtered_df
    
    def write_results(
        self, 
        original_file: Path, 
        results: List[Dict[str, Any]], 
        output_file: Path,
        preserve_existing_hts: bool = True
    ) -> None:
        """
        Write processing results to a new Excel file with color-coded TNVED codes.
        
        Color coding based on similarity score:
        - Score = 1.0 (URL match): No color (white)
        - Score >= 0.185: Green background
        - Score < 0.185: Red background
        
        Args:
            original_file: Path to the original Excel file
            results: List of processing results as dictionaries
            output_file: Path where to save the processed file
            preserve_existing_hts: Whether to preserve existing HTS codes (for selective mode)
        """
        logger.info(f"Writing results to: {output_file}")
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            
            # Read the original file
            df_original = pd.read_excel(original_file, engine='openpyxl')
            
            # Create a copy for the output
            df_output = df_original.copy()
            
            # Find existing HTS column if it exists
            existing_hts_col = None
            for col in df_output.columns:
                if "HTS Code" in str(col) or "HTS_Code" in str(col):
                    existing_hts_col = col
                    break
            
            # Add new columns if they don't exist
            if 'TNVED_Code' not in df_output.columns:
                df_output['TNVED_Code'] = ''
            if 'Selection_Reason' not in df_output.columns:
                df_output['Selection_Reason'] = ''
            
            # Track which rows need coloring and their scores
            row_colors = {}  # {row_idx: confidence_score}
            
            # If preserving existing HTS codes and there's an existing column, copy values
            if preserve_existing_hts and existing_hts_col:
                # Copy existing HTS codes to TNVED_Code column where they exist
                mask = df_output[existing_hts_col].notna() & \
                       (df_output[existing_hts_col].astype(str).str.strip() != '')
                df_output.loc[mask, 'TNVED_Code'] = df_output.loc[mask, existing_hts_col]
                df_output.loc[mask, 'Selection_Reason'] = 'Preserved existing HTS code'
            
            # Apply results to the output DataFrame
            for result in results:
                row_idx = result.get('row_index', -1)
                if 0 <= row_idx < len(df_output):
                    # Only overwrite if we don't have an existing value or not preserving
                    if not preserve_existing_hts or pd.isna(df_output.loc[row_idx, 'TNVED_Code']) or \
                       str(df_output.loc[row_idx, 'TNVED_Code']).strip() == '':
                        df_output.loc[row_idx, 'TNVED_Code'] = result.get('tnved_code', '')
                        df_output.loc[row_idx, 'Selection_Reason'] = result.get('selection_reason', '')
                        
                        # Store confidence score for coloring
                        confidence_score = result.get('confidence_score')
                        if confidence_score is not None:
                            row_colors[row_idx] = confidence_score
            
            # Write to Excel file
            df_output.to_excel(output_file, index=False, engine='openpyxl')
            
            # Apply color coding to TNVED_Code column
            self._apply_color_coding(output_file, row_colors)
            
            logger.info(f"Successfully wrote {len(results)} results to {output_file} with color coding")
            
        except Exception as e:
            logger.error(f"Failed to write results: {e}")
            raise
    
    def _apply_color_coding(self, excel_file: Path, row_colors: Dict[int, float]) -> None:
        """
        Apply color coding to TNVED_Code cells based on similarity scores.
        
        Color rules:
        - Score = 1.0 (URL match): No color (white/default)
        - Score >= 0.185: Green background (00FF00)
        - Score < 0.185: Red background (FF0000)
        
        Args:
            excel_file: Path to the Excel file
            row_colors: Dictionary mapping row indices to confidence scores
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            
            # Load the workbook
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Find TNVED_Code column index
            tnved_col_idx = None
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == 'TNVED_Code':
                    tnved_col_idx = idx
                    break
            
            if tnved_col_idx is None:
                logger.warning("TNVED_Code column not found, skipping color coding")
                return
            
            # Define color fills
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Apply colors based on scores
            for row_idx, confidence_score in row_colors.items():
                # Excel rows are 1-indexed and have header, so add 2
                excel_row = row_idx + 2
                cell = ws.cell(row=excel_row, column=tnved_col_idx)
                
                # Apply color based on score
                if confidence_score == 1.0:
                    # URL match - no color (leave default)
                    pass
                elif confidence_score >= 0.185:
                    # High confidence - green
                    cell.fill = green_fill
                else:
                    # Low confidence - red
                    cell.fill = red_fill
            
            # Save the workbook
            wb.save(excel_file)
            logger.debug(f"Applied color coding to {len(row_colors)} cells")
            
        except Exception as e:
            logger.warning(f"Failed to apply color coding: {e}")
            # Don't raise - color coding is optional enhancement
    
    def get_processing_statistics(
        self, 
        file_path: Path, 
        process_mode: str = "all"
    ) -> Dict[str, int]:
        """
        Get statistics about what will be processed without actually processing.
        
        Args:
            file_path: Path to the Excel file
            process_mode: "all" or "empty_only"
            
        Returns:
            Dictionary with processing statistics
        """
        logger.info(f"Getting processing statistics for: {file_path}")
        
        try:
            validation = self.validate_file(file_path)
            if not validation.is_valid:
                return {
                    'total_rows': 0,
                    'rows_to_process': 0,
                    'rows_to_skip': 0,
                    'empty_descriptions': 0,
                    'existing_hts_codes': 0,
                    'skipped_with_existing_codes': 0
                }
            
            df = pd.read_excel(file_path, engine='openpyxl')
            total_rows = len(df)
            
            # Find columns
            description_col = next(
                col for col in df.columns 
                if "Product Detailed Description" in str(col)
            )
            
            hts_col = None
            for col in df.columns:
                if "HTS Code" in str(col) or "HTS_Code" in str(col):
                    hts_col = col
                    break
            
            # Count empty descriptions
            empty_descriptions = df[description_col].isna().sum() + \
                               (df[description_col].astype(str).str.strip() == '').sum()
            
            # Count existing HTS codes
            existing_hts_codes = 0
            if hts_col:
                existing_hts_codes = (~df[hts_col].isna()).sum() - \
                                   (df[hts_col].astype(str).str.strip() == '').sum()
            
            # Calculate rows to process based on mode
            skipped_with_existing_codes = 0
            if process_mode == "all":
                rows_to_process = total_rows - empty_descriptions
                rows_to_skip = empty_descriptions
            else:  # empty_only
                if hts_col:
                    # Only process rows with descriptions but no HTS codes
                    has_description = ~(df[description_col].isna() | 
                                      (df[description_col].astype(str).str.strip() == ''))
                    has_no_hts = df[hts_col].isna() | (df[hts_col].astype(str).str.strip() == '')
                    has_existing_hts = ~has_no_hts
                    
                    rows_to_process = (has_description & has_no_hts).sum()
                    skipped_with_existing_codes = (has_description & has_existing_hts).sum()
                    rows_to_skip = total_rows - rows_to_process
                else:
                    # No HTS column, same as "all" mode
                    rows_to_process = total_rows - empty_descriptions
                    rows_to_skip = empty_descriptions
            
            stats = {
                'total_rows': total_rows,
                'rows_to_process': rows_to_process,
                'rows_to_skip': rows_to_skip,
                'empty_descriptions': empty_descriptions,
                'existing_hts_codes': existing_hts_codes,
                'skipped_with_existing_codes': skipped_with_existing_codes
            }
            
            logger.info(f"Processing statistics: {stats}")
            return stats
            
        except Exception as e:
            logger.error(f"Failed to get processing statistics: {e}")
            return {
                'total_rows': 0,
                'rows_to_process': 0,
                'rows_to_skip': 0,
                'empty_descriptions': 0,
                'existing_hts_codes': 0,
                'skipped_with_existing_codes': 0
            }
    
    def generate_completion_summary(
        self, 
        file_path: Path, 
        process_mode: str,
        processed_count: int,
        error_count: int
    ) -> Dict[str, Any]:
        """
        Generate a completion summary for selective processing mode reporting.
        
        Args:
            file_path: Path to the original Excel file
            process_mode: "all" or "empty_only"
            processed_count: Number of rows actually processed
            error_count: Number of processing errors
            
        Returns:
            Dictionary with completion summary
        """
        logger.info(f"Generating completion summary for: {file_path}")
        
        try:
            stats = self.get_processing_statistics(file_path, process_mode)
            
            summary = {
                'process_mode': process_mode,
                'total_rows': stats['total_rows'],
                'rows_processed': processed_count,
                'rows_skipped': stats['rows_to_skip'],
                'successful_processing': processed_count - error_count,
                'processing_errors': error_count,
                'empty_descriptions': stats['empty_descriptions'],
                'existing_hts_codes': stats['existing_hts_codes']
            }
            
            if process_mode == "empty_only":
                summary['skipped_with_existing_codes'] = stats['skipped_with_existing_codes']
                summary['reason_for_skipping'] = "Rows with existing HTS codes were preserved"
            else:
                summary['reason_for_skipping'] = "Only rows with empty descriptions were skipped"
            
            # Generate human-readable summary message
            if process_mode == "empty_only":
                summary['message'] = (
                    f"Processed {processed_count} rows out of {stats['total_rows']} total rows. "
                    f"Skipped {stats['skipped_with_existing_codes']} rows with existing HTS codes "
                    f"and {stats['empty_descriptions']} rows with empty descriptions. "
                    f"Successfully processed {processed_count - error_count} rows, "
                    f"with {error_count} errors."
                )
            else:
                summary['message'] = (
                    f"Processed {processed_count} rows out of {stats['total_rows']} total rows. "
                    f"Skipped {stats['empty_descriptions']} rows with empty descriptions. "
                    f"Successfully processed {processed_count - error_count} rows, "
                    f"with {error_count} errors."
                )
            
            logger.info(f"Completion summary: {summary['message']}")
            return summary
            
        except Exception as e:
            logger.error(f"Failed to generate completion summary: {e}")
            return {
                'process_mode': process_mode,
                'total_rows': 0,
                'rows_processed': processed_count,
                'rows_skipped': 0,
                'successful_processing': processed_count - error_count,
                'processing_errors': error_count,
                'message': f"Processing completed with {processed_count} rows processed and {error_count} errors."
            }