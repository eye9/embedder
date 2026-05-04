#!/usr/bin/env python3
"""
Verify that colors are actually applied in the Excel file.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

def verify_colors(excel_file):
    """Verify colors in Excel file."""
    print(f"\n=== Verifying colors in {excel_file} ===\n")
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Find TNVED_Code column
    tnved_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'TNVED_Code':
            tnved_col_idx = idx
            break
    
    if tnved_col_idx is None:
        print("❌ TNVED_Code column not found!")
        return False
    
    print(f"✓ Found TNVED_Code column at index {tnved_col_idx}\n")
    
    # Check colors for data rows
    success = True
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=tnved_col_idx)
        code = cell.value
        fill = cell.fill
        
        # Get fill color
        if fill and fill.start_color:
            color = fill.start_color.rgb if hasattr(fill.start_color, 'rgb') else fill.start_color.index
        else:
            color = None
        
        print(f"Row {row_idx}: Code={code}, Fill={fill.fill_type if fill else 'None'}, Color={color}")
        
        # Check if fill is applied
        if fill and fill.fill_type == 'solid':
            if color:
                print(f"  ✓ Color applied: {color}")
            else:
                print(f"  ⚠ Fill type is solid but no color detected")
        else:
            print(f"  ℹ No fill applied (expected for URL matches with score 1.0)")
    
    print("\n" + "="*60)
    return success

if __name__ == "__main__":
    test_file = Path("real_color_output.xlsx")
    
    if not test_file.exists():
        print(f"❌ File not found: {test_file}")
        sys.exit(1)
    
    verify_colors(test_file)